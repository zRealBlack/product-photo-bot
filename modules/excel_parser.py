"""
excel_parser.py
Parses the Excel file and returns a structured list of products.

Excel structure:
- Multiple sheets (e.g. "الالكترونيات", "الصحة والجمال")
- Each sheet has section header rows (rows with no serial code, usually colored)
- Data rows have: serial_code, category, brand, model_name, qty, price
"""

import openpyxl
from openpyxl.styles import PatternFill


def is_section_header(row_values: list) -> bool:
    """
    Detect section header rows: no serial code in col A,
    but has text in col B or D, or the row is fully merged.
    """
    serial = str(row_values[0]).strip() if row_values[0] else ""
    model = str(row_values[3]).strip() if len(row_values) > 3 and row_values[3] else ""
    # Section headers have no serial code but have descriptive text
    if not serial and model:
        return True
    return False


def parse_excel(file_path: str) -> dict:
    """
    Parse the Excel file and return a dict with:
    {
        "excel_name": str,
        "products": [
            {
                "serial_code": str,
                "category": str,
                "brand": str,
                "model_name": str,
                "section_name": str,
                "sheet_name": str,
            },
            ...
        ]
    }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    import os as _os
    excel_name = _os.path.splitext(_os.path.basename(file_path))[0]
    products = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        current_section = sheet_name  # fallback section name

        for row in ws.iter_rows(values_only=True):
            row_values = list(row)

            # Skip fully empty rows
            if all((v is None or str(v).strip() == "") for v in row_values):
                continue

            # Get all non-empty cell values for this row
            non_empty = [str(v).strip() for v in row_values if v is not None and str(v).strip() != ""]

            serial = str(row_values[0]).strip() if row_values[0] else ""
            category = str(row_values[1]).strip() if len(row_values) > 1 and row_values[1] else ""
            brand = str(row_values[2]).strip() if len(row_values) > 2 and row_values[2] else ""
            model = str(row_values[3]).strip() if len(row_values) > 3 and row_values[3] else ""

            # Skip header row (contains column labels)
            if serial in ("كود الصنف", "Serial", "Code", "الكود", "التصنيف", "الصنف"):
                continue

            # Detect section header rows
            # Case 1: Only exactly ONE cell in the entire row has text (perfect for merged rows!)
            if len(non_empty) == 1:
                val = non_empty[0]
                if val not in ("كود الصنف", "Serial", "Code", "الكود", "التصنيف", "الصنف"):
                    current_section = val
                continue
                
            # Case 2: Row has no serial code (Col A is empty) but has other text (e.g. text in Col B/C)
            if not serial and len(non_empty) > 0:
                val = " - ".join(non_empty)
                if val not in ("كود الصنف", "Serial", "Code", "الكود", "التصنيف", "الصنف"):
                    current_section = val
                continue

            # Valid product row: must have a serial code and model name
            if serial and model:
                products.append({
                    "serial_code": serial,
                    "category": category,
                    "brand": brand,
                    "model_name": model,
                    "section_name": current_section,
                    "sheet_name": sheet_name,
                })

    return {
        "excel_name": excel_name,
        "products": products,
    }
